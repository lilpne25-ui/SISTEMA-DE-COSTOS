from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .models import LlmBomRow
from .normalization import normalize_header, to_text

ERP_HEADERS = [
    "PartNo",
    "Revision",
    "Description",
    "AltDescription1",
    "AltDescription2",
    "DescExtra",
    "Quantity",
    "IssueUM",
    "ConsumptionConv",
    "UM",
    "Cost",
    "Source",
    "Drawing",
    "Leadtime",
    "Level",
    "Location",
    "Memo1",
    "Memo2",
    "Parent",
    "Productline",
    "Sequence",
    "ShotCode",
    "Tag",
    "Category",
    "fabricante",
    "filepat",
    "Material",
    "Peso",
    "largo",
    "ancho",
    "espesor",
    "tratamiento",
]


def write_output_xlsx(path: str, rows: Iterable[LlmBomRow], *, template_path: str | None) -> None:
    wb: Workbook
    ws = None

    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
        ws = pick_template_sheet(wb)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "GLOBAL_SHOP"

    assert ws is not None
    headers = [to_text(v) for v in ws[1]]
    expected = [normalize_header(h) for h in ERP_HEADERS]
    current = [normalize_header(h) for h in headers]
    if current[: len(expected)] != expected:
        ws.delete_rows(1, ws.max_row)
        ws.append(ERP_HEADERS)
    else:
        clear_sheet_data(ws, start_row=2)

    for row in rows:
        ws.append(row.erp_row)

    ws.freeze_panes = "A2"
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def pick_template_sheet(wb: Workbook):
    expected = normalize_header("PartNo")
    for ws in wb.worksheets:
        first = normalize_header(to_text(ws.cell(1, 1).value))
        if first == expected:
            return ws
    return wb.active


def clear_sheet_data(ws, *, start_row: int) -> None:
    if ws.max_row < start_row:
        return
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
