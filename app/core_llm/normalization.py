from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "item": ("item", "line", "linea", "pos", "position"),
    "part_no": ("part no", "part number", "partno", "numero parte", "no parte", "codigo"),
    "description": ("description", "descripcion", "desc"),
    "revision": ("revision", "rev"),
    "alt_description_1": ("altdescription1", "alt description 1", "descripcion 2", "descripcion2"),
    "alt_description_2": ("altdescription2", "alt description 2", "descripcion 3", "descripcion3"),
    "desc_extra": ("descextra", "desc extra", "descripcion extra"),
    "material": ("material", "grade", "alloy", "tipo material"),
    "shape": ("shape", "forma", "profile", "perfil", "tipo"),
    "part_type": ("part type", "tipo parte", "tipoparte"),
    "qty": ("qty", "quantity", "cantidad", "cant"),
    "stock_size": ("stock size", "stocksize", "raw stock", "size", "medida", "dimension", "dims"),
    "length": ("length", "largo", "longitud", "len"),
    "diameter": ("diameter", "diametro", "diam"),
    "width": ("width", "ancho"),
    "thickness": ("thickness", "espesor"),
    "thickness_or_diameter": ("espesor diametro", "espesor-diametro", "espesor di?metro", "espesor-di?metro"),
    "weight": ("weight", "peso"),
    "issue_um": ("issueum", "issue um", "um emision"),
    "consumption_conv": ("consumptionconv", "consumption conv", "factor consumo", "conv consumo"),
    "um": ("uom", "unidad", "unidad medida"),
    "source": ("source", "origen"),
    "drawing": ("drawing", "dibujo"),
    "leadtime": ("leadtime", "lead time", "tiempo entrega"),
    "level": ("level", "nivel"),
    "location": ("location", "ubicacion"),
    "memo1": ("memo1", "nota1"),
    "memo2": ("memo2", "nota2"),
    "parent": ("parent", "padre", "ensamble", "assembly"),
    "productline": ("productline", "product line", "linea producto"),
    "sequence": ("sequence", "secuencia"),
    "shotcode": ("shotcode", "shot code"),
    "tag": ("tag", "etiqueta"),
    "category": ("category", "categoria"),
    "fabricante": ("fabricante", "manufacturer", "maker"),
    "fabricado_por": ("fabricadopor", "fabricado por"),
    "path": ("ruta", "path", "filepath", "file path"),
    "thermal_treatment": ("tratamiento termico", "tratamiento t?rmico", "tratamiento"),
}


def read_source_rows(source_path: str) -> tuple[list[dict[str, object]], str]:
    wb = load_workbook(filename=source_path, data_only=True)
    ws = wb.active
    sheet_name = ws.title

    values = list(ws.iter_rows(values_only=True))
    if not values:
        wb.close()
        raise ValueError("El archivo no contiene datos.")

    header_row_idx, header_map = detect_header(values)
    if header_row_idx < 0 or not header_map:
        wb.close()
        raise ValueError("No se pudieron detectar columnas utiles del Excel de SOLIDWORKS.")

    rows: list[dict[str, object]] = []
    for row_idx in range(header_row_idx + 1, len(values)):
        raw_row = values[row_idx]
        if row_is_empty(raw_row):
            continue
        mapped: dict[str, str] = {}
        for key, col_idx in header_map.items():
            mapped[key] = to_text(raw_row[col_idx] if col_idx < len(raw_row) else None)
        rows.append({"source_row": row_idx + 1, "mapped": mapped})

    wb.close()
    return rows, sheet_name


def detect_header(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    best_row = -1
    best_score = 0
    best_map: dict[str, int] = {}
    probe_rows = min(30, len(rows))
    for idx in range(probe_rows):
        row = rows[idx]
        header_map: dict[str, int] = {}
        for col_idx, value in enumerate(row):
            cell = normalize_header(to_text(value))
            if not cell:
                continue
            for key, aliases in HEADER_ALIASES.items():
                for alias in aliases:
                    if header_alias_match(cell, alias):
                        header_map.setdefault(key, col_idx)
                        break
        score = len(header_map)
        if score > best_score:
            best_score = score
            best_row = idx
            best_map = header_map
    return best_row, best_map


def infer_parent_base(rows: list[dict[str, object]], source_path: str) -> str:
    for row in rows:
        mapped = row["mapped"]
        if isinstance(mapped, dict):
            part = (mapped.get("part_no") or "").strip()
            if part:
                return part
    return Path(source_path).stem


def extract_numbers(text: str) -> list[float]:
    normalized = text.replace(",", ".")
    return [float(x) for x in re.findall(r"[-+]?\d*\.?\d+", normalized)]


def row_is_empty(row: tuple[object, ...]) -> bool:
    for value in row:
        if to_text(value).strip():
            return False
    return True


def to_float(value: object) -> float | None:
    if value is None:
        return None
    text = to_text(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    match = re.search(r"[-+]?\d*\.?\d+", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(text: str) -> str:
    folded = unicodedata.normalize("NFKD", text)
    ascii_only = "".join(ch for ch in folded if not unicodedata.combining(ch))
    cleaned = re.sub(r"[^a-z0-9\s]+", " ", ascii_only.lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def header_alias_match(cell: str, alias: str) -> bool:
    alias_norm = normalize_header(alias)
    if not alias_norm:
        return False
    if cell == alias_norm:
        return True
    if cell.replace(" ", "") == alias_norm.replace(" ", ""):
        return True
    pattern = rf"(?<![a-z0-9]){re.escape(alias_norm)}(?![a-z0-9])"
    return re.search(pattern, cell) is not None
